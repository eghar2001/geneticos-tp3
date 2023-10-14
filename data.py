from copy import deepcopy

import openpyxl
from ciudad import Ciudad



excel_dataframe = openpyxl.load_workbook("TablaCapitales.xlsx")

dataframe = excel_dataframe.active

def get_ciudades():
    """Funcion que retorna una lista con todas las ciudades disponibles"""
    ubicaciones = [[-34.6134525,-58.4140872], [-31.4016247,-64.2161879], [-27.4852401,-58.8241493],[-26.1716109,-58.1958289],
                   [-34.9178973,-57.9703371], [-29.4133186,-66.8624726], [-32.8840183,-68.8671882], [-38.9479696,-68.1356575],
                   [-31.7472389,-60.5561626], [-27.3962271,-55.9657153], [-43.2984093,-65.1068251], [-27.4584468,-59.0010108],
                   [-51.6263882,-69.3129485], [-28.4644524,-65.7894085], [-26.8280277,-65.2421638], [-24.2007152,-65.3215206],
                   [-24.7946597,-65.4663374], [-31.5326152,-68.5450365], [-33.2977218,-66.3398788], [-31.6161908,-60.6861045],
                   [-36.6186225,-64.3159992], [-27.7992032,-64.2900517], [-54.8110607,-68.3225449],[-40.8249897,-63.0105253]
                   ]
    Ciudad._last_id = 0
    cant_columnas = dataframe.max_column - 2
    ciudades = []
    id = 0
    for (col,ubi) in zip(dataframe.iter_cols(2,cant_columnas),ubicaciones):
        id+=1
        ciudades.append(Ciudad(id, col[0].value,ubi))
    return ciudades



def get_distancias():
    """Funcion que retorna un conjunto de distancias entre distintas ciudades
        Retorna un set conformado por tuplas (ciu_1, ciu_2, dist)
    """
    cant_filas = dataframe.max_row - 2
    cant_columnas = dataframe.max_column - 2
    distancias = []
    ciudades = get_ciudades()
    fila_inicial = 3
    for ind_col in range(2,cant_columnas ):
        ciu_act = ciudades[ind_col-2]
        for ind_row in range(fila_inicial,cant_filas+1):
            ciu_dist = ciudades[ind_row-2]
            dist = int(dataframe.cell(column=ind_col, row=ind_row).value)
            distancias.append((ciu_act.id, ciu_dist.id, dist))
        fila_inicial += 1
    return distancias

ciudades = get_ciudades()
distancias_entre_ciudades = get_distancias()